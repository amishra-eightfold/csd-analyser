"""Excel export module for CSD Analyzer."""

from typing import Dict, List, Any, Optional
import pandas as pd
import numpy as np
from datetime import datetime
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from ..analysis.ticket_analysis import TicketAnalyzer
from ..analysis.pattern_analysis import PatternAnalyzer

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Exports support data analysis to Excel format."""
    
    def __init__(self, df: pd.DataFrame):
        """
        Initialize Excel exporter.
        
        Args:
            df (pd.DataFrame): DataFrame containing support ticket data
        """
        self.df = df
        self.ticket_analyzer = TicketAnalyzer(df)
        self.pattern_analyzer = PatternAnalyzer(df)
        
        # Define styles
        self.header_fill = PatternFill(
            start_color='1F4E78',
            end_color='1F4E78',
            fill_type='solid'
        )
        self.header_font = Font(
            color='FFFFFF',
            bold=True,
            size=12
        )
        self.cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def export_analysis(self, 
                       filename: str,
                       customers: Optional[List[str]] = None) -> str:
        """
        Export complete analysis to Excel file.
        
        Args:
            filename (str): Output filename
            customers (Optional[List[str]]): List of customers to include
            
        Returns:
            str: Path to generated Excel file
        """
        try:
            # Create workbook
            wb = Workbook()
            
            # Generate sheets
            self._create_summary_sheet(wb)
            self._create_metrics_sheet(wb)
            self._create_trends_sheet(wb)
            self._create_patterns_sheet(wb)
            self._create_details_sheet(wb)
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Save workbook
            wb.save(filename)
            return filename
            
        except Exception as e:
            logger.error(f"Failed to export analysis to Excel: {str(e)}")
            raise
            
    def _create_summary_sheet(self, wb: Workbook):
        """Create executive summary sheet."""
        try:
            ws = wb.create_sheet("Executive Summary")
            
            # Add title
            ws['A1'] = "Support Analysis Executive Summary"
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:E1')
            
            # Add basic metrics
            metrics = self.ticket_analyzer.get_basic_metrics()
            
            headers = [
                "Metric", "Value", "Previous Period", "Change", "Status"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.border = self.cell_border
            
            # Add metric rows
            metric_rows = [
                ["Total Tickets", metrics['total_tickets']],
                ["Average Resolution Time", f"{metrics['avg_resolution_time']:.1f} days"],
                ["CSAT Score", f"{metrics['avg_csat']:.2f}"],
                ["Escalation Rate", f"{metrics['escalation_rate']:.1%}"]
            ]
            
            for row, metric in enumerate(metric_rows, 4):
                for col, value in enumerate(metric, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    cell.border = self.cell_border
            
            # Add trends summary
            ws['A8'] = "Key Trends"
            ws['A8'].font = Font(size=12, bold=True)
            
            trends = self.pattern_analyzer.analyze_trends()
            
            # Convert trend data into rows
            trend_rows = []
            
            # Volume trends
            if 'volume_trends' in trends:
                trend_rows.append(["Ticket Volume", trends['volume_trends']['trend']])
                trend_rows.append(["  Peak Period", trends['volume_trends'].get('peak_period', 'N/A')])
                trend_rows.append(["  Growth Rate", f"{trends['volume_trends'].get('growth_rate', 0):.1%}"])
            
            # Priority trends
            if 'priority_trends' in trends:
                for priority, data in trends['priority_trends'].items():
                    trend_rows.append([f"Priority {priority}", data['trend']])
                    trend_rows.append([f"  Volume Change", f"{data.get('volume_change', 0):.1%}"])
            
            # Resolution time trends
            if 'resolution_trends' in trends:
                trend_rows.append(["Resolution Time", trends['resolution_trends']['trend']])
                trend_rows.append(["  Improvement", f"{trends['resolution_trends'].get('improvement_rate', 0):.1%}"])
            
            # CSAT trends
            if 'csat_trends' in trends:
                trend_rows.append(["CSAT", trends['csat_trends']['trend']])
                trend_rows.append(["  Change", f"{trends['csat_trends'].get('change', 0):.1%}"])
            
            for row, trend in enumerate(trend_rows, 9):
                for col, value in enumerate(trend, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    cell.border = self.cell_border
            
            # Add critical issues
            ws['A12'] = "Critical Issues"
            ws['A12'].font = Font(size=12, bold=True)
            
            critical_issues = self.ticket_analyzer.get_critical_issues()
            
            if critical_issues:
                headers = [
                    "Case Number", "Subject", "Priority", "Status",
                    "Resolution Time (Days)"
                ]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=13, column=col)
                    cell.value = header
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                    cell.border = self.cell_border
                
                for row, issue in enumerate(critical_issues[:5], 14):  # Show top 5
                    values = [
                        issue['case_number'],
                        issue['subject'],
                        issue['priority'],
                        issue['status'],
                        issue.get('resolution_time', 'N/A')
                    ]
                    
                    for col, value in enumerate(values, 1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = value
                        cell.border = self.cell_border
            
            # Adjust column widths
            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
        except Exception as e:
            logger.error(f"Failed to create summary sheet: {str(e)}")
            raise
            
    def _create_metrics_sheet(self, wb: Workbook):
        """Create detailed metrics sheet."""
        try:
            ws = wb.create_sheet("Detailed Metrics")
            
            # Add title
            ws['A1'] = "Support Metrics Analysis"
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:E1')
            
            # Add status breakdown
            ws['A3'] = "Status Distribution"
            ws['A3'].font = Font(size=12, bold=True)
            
            metrics = self.ticket_analyzer.get_basic_metrics()
            status_data = metrics.get('status_breakdown', {})
            
            headers = ["Status", "Count", "Percentage"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.border = self.cell_border
            
            total = sum(status_data.values())
            
            for row, (status, count) in enumerate(status_data.items(), 5):
                values = [
                    status,
                    count,
                    f"{count/total:.1%}" if total > 0 else "0%"
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    cell.border = self.cell_border
            
            # Add priority breakdown
            ws['A8'] = "Priority Distribution"
            ws['A8'].font = Font(size=12, bold=True)
            
            priority_data = metrics.get('priority_breakdown', {})
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=9, column=col)
                cell.value = header
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.border = self.cell_border
            
            total = sum(priority_data.values())
            
            for row, (priority, count) in enumerate(priority_data.items(), 10):
                values = [
                    priority,
                    count,
                    f"{count/total:.1%}" if total > 0 else "0%"
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    cell.border = self.cell_border
            
            # Add charts
            self._add_pie_chart(
                ws,
                "Status Distribution",
                'A4',
                f'A{4+len(status_data)-1}',
                'B4',
                f'B{4+len(status_data)-1}',
                'E4'
            )
            
            self._add_pie_chart(
                ws,
                "Priority Distribution",
                'A10',
                f'A{10+len(priority_data)-1}',
                'B10',
                f'B{10+len(priority_data)-1}',
                'E10'
            )
            
            # Adjust column widths
            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
        except Exception as e:
            logger.error(f"Failed to create metrics sheet: {str(e)}")
            raise
            
    def _create_trends_sheet(self, wb: Workbook):
        """Create trends analysis sheet."""
        try:
            ws = wb.create_sheet("Trends Analysis")
            
            # Add title
            ws['A1'] = "Support Trends Analysis"
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:E1')
            
            # Add monthly trends
            ws['A3'] = "Monthly Trends"
            ws['A3'].font = Font(size=12, bold=True)
            
            headers = [
                "Month", "Ticket Count", "Avg Resolution Time",
                "Avg CSAT", "Escalation Rate"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.border = self.cell_border
            
            # Calculate monthly metrics
            monthly_data = self.df.resample('ME', on='Created Date').agg({
                'CaseNumber': 'count',
                'Resolution Time (Days)': 'mean',
                'CSAT': 'mean',
                'IsEscalated': 'mean'
            }).reset_index()
            
            for row, (_, data) in enumerate(monthly_data.iterrows(), 5):
                values = [
                    data['Created Date'].strftime('%Y-%m'),
                    data['CaseNumber'],
                    f"{data['Resolution Time (Days)']:.1f}",
                    f"{data['CSAT']:.2f}",
                    f"{data['IsEscalated']:.1%}"
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    cell.border = self.cell_border
            
            # Add charts
            self._add_line_chart(
                ws,
                "Ticket Volume Trend",
                'A5',
                f'A{5+len(monthly_data)-1}',
                'B5',
                f'B{5+len(monthly_data)-1}',
                'G4'
            )
            
            self._add_line_chart(
                ws,
                "Resolution Time Trend",
                'A5',
                f'A{5+len(monthly_data)-1}',
                'C5',
                f'C{5+len(monthly_data)-1}',
                'G20'
            )
            
            # Adjust column widths
            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
        except Exception as e:
            logger.error(f"Failed to create trends sheet: {str(e)}")
            raise
            
    def _create_patterns_sheet(self, wb: Workbook):
        """Create patterns analysis sheet."""
        try:
            ws = wb.create_sheet("Patterns Analysis")
            
            # Add title
            ws['A1'] = "Support Patterns Analysis"
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:E1')
            
            # Add time patterns
            patterns = self.pattern_analyzer.analyze_time_patterns()
            
            # Daily patterns
            ws['A3'] = "Daily Patterns"
            ws['A3'].font = Font(size=12, bold=True)
            
            daily = patterns.get('daily', {})
            if daily:
                headers = ["Hour", "Ticket Count"]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=4, column=col)
                    cell.value = header
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                    cell.border = self.cell_border
                
                distribution = daily.get('distribution', {})
                for row, (hour, count) in enumerate(distribution.items(), 5):
                    values = [hour, count]
                    
                    for col, value in enumerate(values, 1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = value
                        cell.border = self.cell_border
                
                self._add_bar_chart(
                    ws,
                    "Hourly Distribution",
                    'A5',
                    f'A{5+len(distribution)-1}',
                    'B5',
                    f'B{5+len(distribution)-1}',
                    'D4'
                )
            
            # Weekly patterns
            ws['A20'] = "Weekly Patterns"
            ws['A20'].font = Font(size=12, bold=True)
            
            weekly = patterns.get('weekly', {})
            if weekly:
                headers = ["Day", "Ticket Count"]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=21, column=col)
                    cell.value = header
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                    cell.border = self.cell_border
                
                distribution = weekly.get('distribution', {})
                for row, (day, count) in enumerate(distribution.items(), 22):
                    values = [day, count]
                    
                    for col, value in enumerate(values, 1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = value
                        cell.border = self.cell_border
                
                self._add_bar_chart(
                    ws,
                    "Daily Distribution",
                    'A22',
                    f'A{22+len(distribution)-1}',
                    'B22',
                    f'B{22+len(distribution)-1}',
                    'D21'
                )
            
            # Adjust column widths
            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
        except Exception as e:
            logger.error(f"Failed to create patterns sheet: {str(e)}")
            raise
            
    def _create_details_sheet(self, wb: Workbook):
        """Create detailed data sheet."""
        try:
            ws = wb.create_sheet("Detailed Data")
            
            # Add title
            ws['A1'] = "Support Ticket Details"
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:E1')
            
            # Export DataFrame
            headers = [
                "CaseNumber", "Subject", "Status", "Priority",
                "Created Date", "Closed Date", "Resolution Time (Days)",
                "Product Area", "CSAT"
            ]
            
            display_headers = [
                "Case Number", "Subject", "Status", "Priority",
                "Created Date", "Closed Date", "Resolution Time (Days)",
                "Product Area", "CSAT"
            ]
            
            for col, header in enumerate(display_headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.border = self.cell_border
            
            data = self.df[headers].fillna('')
            
            for row, ticket in enumerate(data.values, 4):
                for col, value in enumerate(ticket, 1):
                    cell = ws.cell(row=row, column=col)
                    
                    if isinstance(value, datetime):
                        cell.value = value.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        cell.value = value
                        
                    cell.border = self.cell_border
            
            # Adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Enable filtering
            ws.auto_filter.ref = f"A3:I{len(data) + 3}"
            
        except Exception as e:
            logger.error(f"Failed to create details sheet: {str(e)}")
            raise
            
    def _add_pie_chart(self,
                      ws: Any,
                      title: str,
                      cat_start: str,
                      cat_end: str,
                      val_start: str,
                      val_end: str,
                      pos: str):
        """Add pie chart to worksheet."""
        try:
            chart = PieChart()
            chart.title = title
            
            # Convert cell references to row/col indices
            cat_start_col = ord(cat_start[0]) - ord('A') + 1
            cat_start_row = int(cat_start[1:])
            cat_end_col = ord(cat_end[0]) - ord('A') + 1
            cat_end_row = int(cat_end[1:])
            
            val_start_col = ord(val_start[0]) - ord('A') + 1
            val_start_row = int(val_start[1:])
            val_end_col = ord(val_end[0]) - ord('A') + 1
            val_end_row = int(val_end[1:])
            
            cats = Reference(ws, min_col=cat_start_col, min_row=cat_start_row,
                           max_col=cat_end_col, max_row=cat_end_row)
            vals = Reference(ws, min_col=val_start_col, min_row=val_start_row,
                           max_col=val_end_col, max_row=val_end_row)
            
            chart.add_data(vals)
            chart.set_categories(cats)
            
            ws.add_chart(chart, pos)
            
        except Exception as e:
            logger.error(f"Failed to add pie chart: {str(e)}")
            
    def _add_line_chart(self,
                       ws: Any,
                       title: str,
                       cat_start: str,
                       cat_end: str,
                       val_start: str,
                       val_end: str,
                       pos: str):
        """Add line chart to worksheet."""
        try:
            chart = LineChart()
            chart.title = title
            
            # Convert cell references to row/col indices
            cat_start_col = ord(cat_start[0]) - ord('A') + 1
            cat_start_row = int(cat_start[1:])
            cat_end_col = ord(cat_end[0]) - ord('A') + 1
            cat_end_row = int(cat_end[1:])
            
            val_start_col = ord(val_start[0]) - ord('A') + 1
            val_start_row = int(val_start[1:])
            val_end_col = ord(val_end[0]) - ord('A') + 1
            val_end_row = int(val_end[1:])
            
            cats = Reference(ws, min_col=cat_start_col, min_row=cat_start_row,
                           max_col=cat_end_col, max_row=cat_end_row)
            vals = Reference(ws, min_col=val_start_col, min_row=val_start_row,
                           max_col=val_end_col, max_row=val_end_row)
            
            chart.add_data(vals)
            chart.set_categories(cats)
            
            ws.add_chart(chart, pos)
            
        except Exception as e:
            logger.error(f"Failed to add line chart: {str(e)}")
            
    def _add_bar_chart(self,
                      ws: Any,
                      title: str,
                      cat_start: str,
                      cat_end: str,
                      val_start: str,
                      val_end: str,
                      pos: str):
        """Add bar chart to worksheet."""
        try:
            chart = BarChart()
            chart.title = title
            
            # Convert cell references to row/col indices
            cat_start_col = ord(cat_start[0]) - ord('A') + 1
            cat_start_row = int(cat_start[1:])
            cat_end_col = ord(cat_end[0]) - ord('A') + 1
            cat_end_row = int(cat_end[1:])
            
            val_start_col = ord(val_start[0]) - ord('A') + 1
            val_start_row = int(val_start[1:])
            val_end_col = ord(val_end[0]) - ord('A') + 1
            val_end_row = int(val_end[1:])
            
            cats = Reference(ws, min_col=cat_start_col, min_row=cat_start_row,
                           max_col=cat_end_col, max_row=cat_end_row)
            vals = Reference(ws, min_col=val_start_col, min_row=val_start_row,
                           max_col=val_end_col, max_row=val_end_row)
            
            chart.add_data(vals)
            chart.set_categories(cats)
            
            ws.add_chart(chart, pos)
            
        except Exception as e:
            logger.error(f"Failed to add bar chart: {str(e)}")
            raise
